import sys
import re
import asyncio
import aiohttp
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import worksheet
from const import INPUT_FILE, URL_TEMPLATE, SHEET_PROPERTIES, MESSAGE_TYPES


def create_or_clear_sheet(wb: Workbook, sheet_name: str, col_names: list) -> worksheet:
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    else:
        wb[sheet_name].delete_rows(1, wb[sheet_name].max_row)
    wb[sheet_name].append(col_names)
    return wb[sheet_name]


def set_message_sign(message: str) -> str:
    mess_type = MESSAGE_TYPES.get('vacancy')
    for sign in mess_type.get('sign'):
        if sign in str(message):
            return mess_type.get('type_str')
    mess_type = MESSAGE_TYPES.get('statistic')
    for sign in mess_type.get('sign'):
        if str(message).startswith(sign):
            return mess_type.get('type_str')
    mess_type = MESSAGE_TYPES.get('service')
    for sign in mess_type.get('sign'):
        if sign in str(message):
            return mess_type.get('type_str')
    return ''


def get_detail_stat(message: str) -> list:
    result = [None, None, None, None, None, None, None]  # 7 элементов
    for index, sign in MESSAGE_TYPES.get('statistic').get('detail_sign').items():
        for item in sign:
            param = re.search(item + r'\s([-+]?\d+[.,]?\d*)', message)  # +/-цифры./,цифры
            if param:
                result[index] = param.groups()[0]
            else:
                param = re.search(item + r'\s?\$(\d+-\d+)', message)  # $цифры-цифры
                if param:
                    result[index], result[index + 1] = param.groups()[0].split('-')
    for i, item in enumerate(result):
        try:
            result[i] = int(item)
        except ValueError:
            try:
                result[i] = float(item)
            except ValueError:
                result[i] = None
    return result


def get_detail_vacancy(message: str) -> list:
    result = [None, None, None, None, None, None, None, None]  # 8 элементов
    message_parts = str(message).split('\n')
    for index, sign in MESSAGE_TYPES.get('vacancy').get('detail_sign').items():
        parse_string = ''
        parse_string = message_parts[0] if index in [0, 1] else parse_string
        parse_string = message_parts[1] if index in [2, 3, 4] else parse_string
        parse_string = message_parts[-2] if index == 6 else parse_string
        parse_string = message_parts[-1] if index == 7 else parse_string
        for item in sign:
            match = None
            if index == 0:
                match = re.search(r'(.*)' + item, parse_string)
                result[index] = match.group(1) if match else parse_string
            if index == 1:
                match = re.search(item + r'(.*)', parse_string)
            if index == 2:
                match = re.search(r'(.*)' + item, parse_string)
            if index == 3:
                match = re.search(r'.*, (\d+) ' + item, parse_string)
                if not match:
                    match = re.search(r'.*, (' + item + r'), .*', parse_string)
                if not match:
                    match = re.search(r', (' + item + r')', parse_string)
            if index == 4:
                match = re.search(r'.*' + item + r'(\d+.*)', parse_string)
            if index == 5:
                result[index] = '\n'.join(message_parts[2:-2])
            if index == 6:
                match = re.search(r'(' + item + r'.*)', parse_string)
            if index == 7:
                match = re.search(item + r' (.*)', parse_string)
            if index in [1, 2, 3, 4, 6, 7]:
                result[index] = match.group(1) if match else result[index]
    return result


async def get_page(url, session):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.text()
        else:
            return f'Error code: {response.status}'


async def get_many_pages(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [get_page(url, session) for url in urls]
        results = await asyncio.gather(*tasks)
        return {url: text for url, text in zip(urls, results)}


async def get_vacancies_content(urls: list) -> dict:
    url_texts = await get_many_pages(urls)
    return url_texts


def vacancy_html_parsing(html: str) -> dict:
    result = {}
    parsing_content = BeautifulSoup(html, 'lxml')
    # получение заголовка вакансии
    items_found = []
    for item in parsing_content.find_all('h1'):
        clean_text = item.decode_contents().replace('\n', '')  # удаление переносов строк
        clean_text = re.sub(r'<.*?>', '', clean_text)  # удаление тегов
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()  # удаление лишних пробелов
        items_found.append(clean_text)
    if items_found:
        result['header'] = '\n'.join(items_found)
    # получение текста вакансии
    items_found.clear()
    for item in parsing_content.find_all('div', class_='mb-4'):
        clean_text = item.decode_contents().replace('<br/>', '\n')  # замена переносов строк
        clean_text = re.sub(r'<.*?>', '', clean_text)  # удаление тегов
        clean_text = re.sub(r'\n\s*\n\s*\n', '\n\n', clean_text)  # удаление лишних переносов строк
        clean_text = re.sub(r'\n\n\s+', '\n\n', clean_text).strip()  # удаление лишних пробелов
        items_found.append(clean_text)
    if items_found:
        result['text'] = '\n\n'.join(items_found)
    # получение дополнительной информации о вакансии
    items_found.clear()
    for item in parsing_content.find_all('div', class_='job-additional-info--item-text'):
        clean_text = item.decode_contents().replace('\n', '')  # удаление переносов строк
        clean_text = re.sub(r'<.*?>', '', clean_text)  # удаление тегов
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()  # удаление лишних пробелов
        items_found.append(clean_text)
    if items_found:
        result['additional_info'] = '\n'.join(items_found)
    return result


def main():
    wb = load_workbook(filename=INPUT_FILE)
    # Проверяем наличие листа с сообщениями
    if SHEET_PROPERTIES.get('messages').get('sheet_name') not in wb.sheetnames:
        print(f'В файле нет листа {SHEET_PROPERTIES.get("messages").get("sheet_name")}')
        sys.exit(1)
    ws_mess = wb[SHEET_PROPERTIES.get('messages').get('sheet_name')]
    # Прописываем заголовки столбцов листа с сообщениями
    for i, item in enumerate(SHEET_PROPERTIES.get('messages').get('col_names'), 1):
        ws_mess.cell(row=1, column=i).value = item
    # Определяем тип каждого сообщения и записываем в столбец D
    report.get(MESSAGE_TYPES.get('messages').get('type_str'))[0] = ws_mess.max_row - 1
    report.get(MESSAGE_TYPES.get('messages').get('type_str'))[1] -= 1
    for row in ws_mess.rows:
        if not row[3].value:
            row[3].value = set_message_sign(row[2].value)
            if row[3].value == MESSAGE_TYPES.get('vacancy').get('type_str'):
                row[4].value = str(row[2].value).count('\n\n') + 1
        if row[3].value:
            report.get(MESSAGE_TYPES.get('messages').get('type_str'))[1] += 1
    # Создаем лист со статистикой или очищаем его если он уже есть
    current_sheet = SHEET_PROPERTIES.get('statistic')
    ws_stat = create_or_clear_sheet(wb, current_sheet.get('sheet_name'), current_sheet.get('col_names'))
    # for i, item in enumerate(SHEET_PROPERTIES.get('statistic').get('col_names'), 1):
    #     ws_stat.cell(row=1, column=i).value = item
    # Получаем детальную статистику и записываем в лист со статистикой
    detail_stat = []
    for row in ws_mess.rows:
        if row[3].value == MESSAGE_TYPES.get('statistic').get('type_str'):
            report.get(MESSAGE_TYPES.get('statistic').get('type_str'))[0] += 1
            detail_stat.extend([row[0].value, row[1].value])
            detail_stat.extend(get_detail_stat(row[2].value))
            ws_stat.append(detail_stat)
            if 'None' not in map(str, detail_stat):
                row[5].value = 'Успешно'
                report.get(MESSAGE_TYPES.get('statistic').get('type_str'))[1] += 1
            detail_stat.clear()
    # Создаем лист со служебными командами или очищаем его если он уже есть
    current_sheet = SHEET_PROPERTIES.get('service')
    ws_serv = create_or_clear_sheet(wb, current_sheet.get('sheet_name'), current_sheet.get('col_names'))
    # Получаем сервисные сообщения и записываем в лист со служебными командами
    service_command = []
    for row in ws_mess.rows:
        if row[3].value == MESSAGE_TYPES.get('service').get('type_str'):
            report.get(MESSAGE_TYPES.get('service').get('type_str'))[0] += 1
            service_command.extend([row[0].value, row[1].value, row[2].value])
            ws_serv.append(service_command)
            row[5].value = 'Успешно'
            report.get(MESSAGE_TYPES.get('service').get('type_str'))[1] += 1
            service_command.clear()
    # Создаем лист с вакансиями или очищаем его если он уже есть
    current_sheet = SHEET_PROPERTIES.get('vacancy')
    ws_vac = create_or_clear_sheet(wb, current_sheet.get('sheet_name'), current_sheet.get('col_names'))
    # Проверяем сообщения на наличие нескольких вакансий и записываем их по отдельности в лист с вакансиями
    # одновременно получая детальную информацию по каждой вакансии
    detail_vac = []
    for row in ws_mess.rows:
        if row[3].value == MESSAGE_TYPES.get('vacancy').get('type_str'):
            for item in row[2].value.split('\n\n'):
                report.get(MESSAGE_TYPES.get('vacancy').get('type_str'))[0] += 1
                detail_vac.extend([row[0].value, row[1].value])
                detail_vac.extend(get_detail_vacancy(item))
                ws_vac.append(detail_vac)
                if 'None' not in map(str, detail_vac[0:6] + detail_vac[8:]):
                    row[5].value = 'Успешно'
                    report.get(MESSAGE_TYPES.get('vacancy').get('type_str'))[1] += 1
                detail_vac.clear()
    # Формируем список URL и получаем HTML-страницы с сайта https://djinni.co/ с текстом вакансии
    vacancies_urls = []
    for row in ws_vac.rows:
        if URL_TEMPLATE in row[8].value:
            vacancies_urls.append(row[8].value)
            report.get(MESSAGE_TYPES.get('vacancy_parsing').get('type_str'))[0] += 1
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    vacancies_html = asyncio.run(get_vacancies_content(vacancies_urls))
    # Парсим HTML-страницы с текстом вакансии и дописываем результат в лист с вакансиями
    for row in ws_vac.rows:
        if vacancies_html.get(row[8].value):
            html_parsing_result = vacancy_html_parsing(vacancies_html.get(row[8].value))
            if html_parsing_result:
                row[10].value = html_parsing_result.get('header')
                row[11].value = html_parsing_result.get('additional_info')
                row[12].value = html_parsing_result.get('text')
                report.get(MESSAGE_TYPES.get('vacancy_parsing').get('type_str'))[1] += 1
            else:
                row[10].value = vacancies_html.get(row[8].value)
    # Сохраняем результаты работы в файл
    wb.save(filename=INPUT_FILE)


if __name__ == '__main__':
    report = {}
    for key, value in MESSAGE_TYPES.items():
        report[value.get('type_str')] = [0, 0]
    main()
    print('\tОтчет:')
    for key, value in report.items():
        percent = 0 if value[1] == 0 else value[1] / value[0] * 100
        detail = '(' + ' / '.join([str(value[1]), str(value[0])]) + ')'
        print(f'{key}:\t{percent:>5.1f} %  {detail:^11}')
